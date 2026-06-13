"""VAT report: preview, Excel export (→ S3 + download), and email to consultant."""

from __future__ import annotations

import io
import os
from datetime import date, datetime
from decimal import Decimal
from typing import List, Optional

import boto3
import openpyxl
from botocore.exceptions import ClientError
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from config import settings as app_settings
from database import get_db
from i18n import get_translations
from models import Document, Expense, Invoice, Lessee, RentalAgreement, Settings, Unit, Property
from routers.auth import get_current_user

router = APIRouter(prefix="/vat-report", tags=["vat"])


# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------

class VATInvoiceLine(BaseModel):
    invoice_number: str
    issue_date: date
    billing_period_start: date
    billing_period_end: date
    lessee_name: str
    unit_label: str
    net_amount: Decimal
    vat_amount: Decimal
    gross_amount: Decimal
    invoice_type: str  # standard / credit_note


class VATExpenseLine(BaseModel):
    expense_date: date
    vendor_name: str
    expense_category: str
    description: str
    net_amount: Decimal
    vat_amount: Decimal
    gross_amount: Decimal


class VATReportResponse(BaseModel):
    period_start: date
    period_end: date
    invoices: List[VATInvoiceLine]
    expenses: List[VATExpenseLine]
    # totals
    invoice_net_total: Decimal
    invoice_vat_total: Decimal
    expense_net_total: Decimal
    expense_vat_total: Decimal
    vat_due: Decimal  # invoice VAT collected - expense VAT reclaimable


class VATExportRequest(BaseModel):
    period_start: date
    period_end: date
    upload_to_documents: bool = True


class VATExportResponse(BaseModel):
    download_url: str
    document_id: Optional[int] = None
    expires_in: int = 900


class VATSendRequest(BaseModel):
    period_start: date
    period_end: date


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _build_report(db: Session, org_id: int, period_start: date, period_end: date) -> VATReportResponse:
    # Invoices whose issue_date falls within the VAT period (all non-draft statuses).
    # issue_date is the tax-point date for VAT purposes.
    raw_invoices = (
        db.query(Invoice)
        .filter(
            Invoice.org_id == org_id,
            Invoice.issue_date >= period_start,
            Invoice.issue_date <= period_end,
            Invoice.invoice_status.in_(["pending", "paid", "overdue", "credited"]),
        )
        .order_by(Invoice.issue_date)
        .all()
    )

    invoice_lines: List[VATInvoiceLine] = []
    for inv in raw_invoices:
        # resolve lessee name and unit
        agr = db.query(RentalAgreement).filter_by(agreement_uuid=inv.agreement_uuid).first()
        lessee_name = "—"
        unit_label = "—"
        if agr:
            unit = db.query(Unit).filter_by(id=agr.unit_id).first() if agr.unit_id else None
            prop_id = unit.property_id if unit else agr.property_id
            prop = db.query(Property).filter_by(id=prop_id).first() if prop_id else None
            lessee = db.query(Lessee).filter_by(lessee_uuid=agr.lessee_uuid).first()
            if unit and prop:
                unit_label = f"{prop.name} / {unit.unit_number}"
            elif prop:
                unit_label = prop.name
            if lessee:
                lessee_name = (
                    lessee.company_legal_name
                    or f"{lessee.first_name or ''} {lessee.last_name or ''}".strip()
                )
        invoice_lines.append(VATInvoiceLine(
            invoice_number=inv.invoice_number,
            issue_date=inv.issue_date,
            billing_period_start=inv.billing_period_start,
            billing_period_end=inv.billing_period_end,
            lessee_name=lessee_name,
            unit_label=unit_label,
            net_amount=inv.net_amount,
            vat_amount=inv.vat_amount,
            gross_amount=inv.gross_amount,
            invoice_type=inv.invoice_type,
        ))

    # Expenses within the period
    raw_expenses = (
        db.query(Expense)
        .filter(
            Expense.org_id == org_id,
            Expense.expense_date >= period_start,
            Expense.expense_date <= period_end,
        )
        .order_by(Expense.expense_date)
        .all()
    )

    expense_lines = [
        VATExpenseLine(
            expense_date=exp.expense_date,
            vendor_name=exp.vendor_name,
            expense_category=exp.expense_category.replace("_", " ").title(),
            description=exp.description,
            net_amount=exp.net_amount,
            vat_amount=exp.vat_amount,
            gross_amount=exp.gross_amount,
        )
        for exp in raw_expenses
    ]

    inv_net = sum((l.net_amount for l in invoice_lines), Decimal("0"))
    inv_vat = sum((l.vat_amount for l in invoice_lines), Decimal("0"))
    exp_net = sum((l.net_amount for l in expense_lines), Decimal("0"))
    exp_vat = sum((l.vat_amount for l in expense_lines), Decimal("0"))

    return VATReportResponse(
        period_start=period_start,
        period_end=period_end,
        invoices=invoice_lines,
        expenses=expense_lines,
        invoice_net_total=inv_net,
        invoice_vat_total=inv_vat,
        expense_net_total=exp_net,
        expense_vat_total=exp_vat,
        vat_due=inv_vat - exp_vat,
    )


def _build_excel(report: VATReportResponse, company_name: str) -> bytes:
    wb = openpyxl.Workbook()

    # ---- Styles ----
    from openpyxl.styles import Alignment, Font, PatternFill, numbers
    DARK_FILL = PatternFill("solid", fgColor="1F2937")
    HEADER_FILL = PatternFill("solid", fgColor="312E81")
    WHITE = Font(color="FFFFFF", bold=True)
    BOLD = Font(bold=True)
    MONEY = '#,##0.00'
    CENTER = Alignment(horizontal="center")

    # ---- Summary sheet ----
    ws_sum = wb.active
    ws_sum.title = "VAT Summary"

    ws_sum.column_dimensions["A"].width = 40
    ws_sum.column_dimensions["B"].width = 18

    ws_sum["A1"] = company_name
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A2"] = f"VAT Report — {report.period_start} to {report.period_end}"
    ws_sum["A2"].font = Font(size=11, color="9CA3AF")
    ws_sum.append([])

    headers = ["Category", "Amount"]
    ws_sum.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws_sum.cell(row=4, column=col)
        cell.fill = HEADER_FILL
        cell.font = WHITE

    rows = [
        ("Invoice Revenue (Net)", report.invoice_net_total),
        ("VAT Collected on Invoices", report.invoice_vat_total),
        ("", ""),
        ("Expense (Net)", report.expense_net_total),
        ("VAT Reclaimable on Expenses", report.expense_vat_total),
        ("", ""),
        ("NET VAT DUE / (REFUNDABLE)", report.vat_due),
    ]
    for label, amount in rows:
        ws_sum.append([label, amount if amount != "" else ""])
        r = ws_sum.max_row
        if label.startswith("NET"):
            ws_sum.cell(r, 1).font = BOLD
            ws_sum.cell(r, 2).font = Font(bold=True, color="EF4444" if report.vat_due >= 0 else "10B981")
        if amount != "":
            ws_sum.cell(r, 2).number_format = MONEY

    # ---- Invoices sheet ----
    ws_inv = wb.create_sheet("Invoices")
    inv_headers = ["Invoice No.", "Type", "Issue Date", "Billing Period Start", "Billing Period End", "Lessee", "Unit", "Net", "VAT", "Gross"]
    ws_inv.append(inv_headers)
    for col, _ in enumerate(inv_headers, 1):
        cell = ws_inv.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = WHITE

    col_widths = [18, 12, 14, 18, 18, 30, 25, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws_inv.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for line in report.invoices:
        ws_inv.append([
            line.invoice_number,
            line.invoice_type.replace("_", " ").title(),
            line.issue_date,
            line.billing_period_start,
            line.billing_period_end,
            line.lessee_name,
            line.unit_label,
            line.net_amount,
            line.vat_amount,
            line.gross_amount,
        ])
        r = ws_inv.max_row
        for col in [8, 9, 10]:
            ws_inv.cell(r, col).number_format = MONEY

    # totals row — col 7=Unit(label), 8=Net, 9=VAT, 10=Gross
    tr = ws_inv.max_row + 1
    ws_inv.cell(tr, 7, "TOTAL").font = BOLD
    ws_inv.cell(tr, 8, report.invoice_net_total).number_format = MONEY
    ws_inv.cell(tr, 8).font = BOLD
    ws_inv.cell(tr, 9, report.invoice_vat_total).number_format = MONEY
    ws_inv.cell(tr, 9).font = BOLD

    # ---- Expenses sheet ----
    ws_exp = wb.create_sheet("Expenses")
    exp_headers = ["Date", "Vendor", "Category", "Description", "Net", "VAT", "Gross"]
    ws_exp.append(exp_headers)
    for col, _ in enumerate(exp_headers, 1):
        cell = ws_exp.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = WHITE

    exp_col_widths = [14, 25, 22, 40, 14, 14, 14]
    for i, w in enumerate(exp_col_widths, 1):
        ws_exp.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for line in report.expenses:
        ws_exp.append([
            line.expense_date,
            line.vendor_name,
            line.expense_category,
            line.description,
            line.net_amount,
            line.vat_amount,
            line.gross_amount,
        ])
        r = ws_exp.max_row
        for col in [5, 6, 7]:
            ws_exp.cell(r, col).number_format = MONEY

    tr = ws_exp.max_row + 1
    ws_exp.cell(tr, 4, "TOTAL").font = BOLD
    ws_exp.cell(tr, 5, report.expense_net_total).number_format = MONEY
    ws_exp.cell(tr, 5).font = BOLD
    ws_exp.cell(tr, 6, report.expense_vat_total).number_format = MONEY
    ws_exp.cell(tr, 6).font = BOLD

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _s3_client():
    return boto3.client(
        "s3",
        region_name=app_settings.AWS_REGION,
        aws_access_key_id=app_settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=app_settings.AWS_SECRET_ACCESS_KEY,
    )


def _upload_to_s3(data: bytes, s3_key: str, mime: str) -> str:
    bucket = app_settings.AWS_S3_BUCKET_NAME
    _s3_client().put_object(Bucket=bucket, Key=s3_key, Body=data, ContentType=mime)
    return bucket


def _presigned_url(s3_key: str, expires: int = 900) -> str:
    return _s3_client().generate_presigned_url(
        "get_object",
        Params={"Bucket": app_settings.AWS_S3_BUCKET_NAME, "Key": s3_key},
        ExpiresIn=expires,
    )


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get("", response_model=VATReportResponse)
def get_vat_report(
    period_start: date = Query(...),
    period_end: date = Query(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
) -> VATReportResponse:
    if period_end < period_start:
        raise HTTPException(status_code=400, detail="period_end must be >= period_start")
    return _build_report(db, current_user.org_id, period_start, period_end)


@router.post("/export", response_model=VATExportResponse)
def export_vat_report(
    payload: VATExportRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
) -> VATExportResponse:
    report = _build_report(db, current_user.org_id, payload.period_start, payload.period_end)

    org_settings = db.query(Settings).filter_by(org_id=current_user.org_id).first()
    company_name = org_settings.company_name if org_settings else "Organisation"

    xlsx_bytes = _build_excel(report, company_name)
    filename = f"VAT-Report-{payload.period_start}-{payload.period_end}.xlsx"
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    import uuid as _uuid
    s3_key = f"orgs/{current_user.org_id}/vat_reports/{_uuid.uuid4()}.xlsx"

    try:
        _upload_to_s3(xlsx_bytes, s3_key, mime)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"S3 upload failed: {exc}")

    doc_id: Optional[int] = None
    if payload.upload_to_documents:
        # Use org_id=0 as a sentinel entity_id for the vat_report entity type
        doc = Document(
            org_id=current_user.org_id,
            display_name=filename,
            s3_bucket=app_settings.AWS_S3_BUCKET_NAME,
            s3_key=s3_key,
            file_size_bytes=len(xlsx_bytes),
            mime_type=mime,
            related_entity_type="vat_report",
            related_entity_id=current_user.org_id,
            document_category="vat_report",
            uploaded_by_user_id=current_user.id,
        )
        db.add(doc)
        db.commit()
        db.refresh(doc)
        doc_id = doc.id

    url = _presigned_url(s3_key, expires=900)
    return VATExportResponse(download_url=url, document_id=doc_id, expires_in=900)


@router.post("/send")
def send_vat_report(
    payload: VATSendRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
) -> dict:
    org_settings = db.query(Settings).filter_by(org_id=current_user.org_id).first()
    if not org_settings or not org_settings.vat_consultant_email:
        raise HTTPException(status_code=400, detail="No VAT consultant email configured in Settings.")

    report = _build_report(db, current_user.org_id, payload.period_start, payload.period_end)
    company_name = org_settings.company_name
    xlsx_bytes = _build_excel(report, company_name)
    filename = f"VAT-Report-{payload.period_start}-{payload.period_end}.xlsx"

    import base64
    import resend

    resend.api_key = app_settings.RESEND_API_KEY

    # ── i18n ─────────────────────────────────────────────────────────────────
    lang = org_settings.invoice_language if org_settings and org_settings.invoice_language else "en"
    t = get_translations(lang)

    to_email = app_settings.DEV_EMAIL or org_settings.vat_consultant_email
    currency = org_settings.currency if org_settings else "EUR"
    vat_id = org_settings.company_vat_number or "—"
    company_address = (org_settings.company_address or "").replace("\n", " · ") if org_settings else ""
    reply_to = org_settings.billing_email_sender if org_settings and org_settings.billing_email_sender else None

    period_start_str = payload.period_start.strftime("%d %b %Y")
    period_end_str   = payload.period_end.strftime("%d %b %Y")

    subject = f"VAT Report {period_start_str} – {period_end_str} | {company_name} | {vat_id}"

    def _fmt(amount: Decimal) -> str:
        return f"{currency} {amount:,.2f}"

    def _vat_row(label: str, value: str, bold: bool = False, highlight: bool = False) -> str:
        weight = "font-weight:700;" if bold else ""
        bg     = "background:#f0fdf4;" if highlight and float(value.split()[-1].replace(",", "")) < 0 else \
                 "background:#fef2f2;" if highlight else ""
        return (
            f'<tr style="{bg}">'
            f'<td style="padding:9px 16px 9px 0;color:#555;font-size:13px;{weight}">{label}</td>'
            f'<td style="padding:9px 0;color:#111;font-size:13px;{weight}text-align:right;">{value}</td>'
            f'</tr>'
        )

    vat_due_positive = report.vat_due >= 0

    summary_rows = (
        _vat_row(t["vat_email_row_inv_net"], _fmt(report.invoice_net_total))
        + _vat_row(t["vat_email_row_inv_vat"], _fmt(report.invoice_vat_total))
        + '<tr><td colspan="2" style="padding:2px 0;border-bottom:1px solid #e8e8e8;"></td></tr>'
        + _vat_row(t["vat_email_row_exp_net"], _fmt(report.expense_net_total))
        + _vat_row(t["vat_email_row_exp_vat"], _fmt(report.expense_vat_total))
        + '<tr><td colspan="2" style="padding:2px 0;border-bottom:2px solid #111;"></td></tr>'
        + _vat_row(t["vat_email_row_vat_due"], _fmt(report.vat_due), bold=True, highlight=True)
    )

    footer_parts = [company_name]
    if company_address:
        footer_parts.append(company_address)
    if org_settings and org_settings.company_vat_number:
        footer_parts.append(f'{t["vat_company_label"]} {org_settings.company_vat_number}')
    footer_text = " &nbsp;·&nbsp; ".join(footer_parts)

    email_html = f"""<!DOCTYPE html>
<html lang="{lang}">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#ffffff;font-family:-apple-system,'Helvetica Neue',Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#ffffff;padding:32px 16px;">
    <tr><td align="center">
      <table width="100%" cellpadding="0" cellspacing="0" style="max-width:560px;background:#ffffff;">

        <!-- Top accent bar -->
        <tr><td style="height:4px;background:#111111;"></td></tr>

        <!-- Body -->
        <tr><td style="padding:36px 40px 28px;">

          <p style="margin:0 0 20px;font-size:15px;color:#111;">{t["vat_email_salutation"]}</p>

          <p style="margin:0 0 16px;font-size:14px;color:#444;line-height:1.7;">
            {t["vat_email_intro"].format(company=company_name, vat_id=vat_id,
              period_start=period_start_str, period_end=period_end_str)}
          </p>

          <p style="margin:0 0 24px;font-size:14px;color:#444;line-height:1.7;">
            {t["vat_email_request"]}
          </p>

          <!-- VAT summary table -->
          <table width="100%" cellpadding="0" cellspacing="0"
                 style="margin-bottom:28px;border:1px solid #e8e8e8;border-radius:6px;overflow:hidden;">
            <thead>
              <tr style="background:#f8f8f8;border-bottom:1px solid #e8e8e8;">
                <th style="padding:10px 16px 10px 0;font-size:11px;font-weight:700;text-transform:uppercase;
                           letter-spacing:0.6px;color:#888;text-align:left;">{t["vat_email_col_position"]}</th>
                <th style="padding:10px 0;font-size:11px;font-weight:700;text-transform:uppercase;
                           letter-spacing:0.6px;color:#888;text-align:right;">{t["vat_email_col_amount"]}</th>
              </tr>
            </thead>
            <tbody>
              {summary_rows}
            </tbody>
          </table>

          <p style="margin:0 0 28px;font-size:14px;color:#444;line-height:1.7;">{t["vat_email_closing"]}</p>

          <p style="margin:0;font-size:14px;color:#444;">
            {t["email_kind_regards"]}<br>
            <strong style="color:#111;">{company_name}</strong>
          </p>

        </td></tr>

        <!-- Footer -->
        <tr><td style="padding:16px 40px;background:#f8f8f8;border-top:1px solid #e8e8e8;
                       font-size:11px;color:#999;text-align:center;line-height:1.8;">
          {footer_text}
        </td></tr>

      </table>
    </td></tr>
  </table>
</body>
</html>"""

    send_kwargs: dict = {
        "from": app_settings.INVOICE_FROM_EMAIL,
        "to": [to_email],
        "subject": subject,
        "html": email_html,
        "attachments": [
            {
                "filename": filename,
                "content": base64.b64encode(xlsx_bytes).decode(),
            }
        ],
    }
    if reply_to:
        send_kwargs["reply_to"] = reply_to
    resend.Emails.send(send_kwargs)

    return {"detail": f"Report sent to {org_settings.vat_consultant_email}"}
